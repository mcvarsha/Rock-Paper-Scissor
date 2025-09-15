from openpyxl import Workbook, load_workbook

# Global variables for Excel workbook and worksheet
wb = Workbook()
ws = None

# Function to initialize Excel workbook and worksheet
def initialize_excel():
    global wb, ws
    try:
        wb = load_workbook('scores.xlsx')  # Load existing workbook
    except FileNotFoundError:
        wb = Workbook()  # Create a new workbook if file doesn't exist
    ws = wb.active
    ws.title = "Scores"
    ws.append(["User Score", "Computer Score"])

# Function to update scores in Excel sheet
def update_scores(user_score, comp_score):
    ws.append([user_score, comp_score])
    wb.save('scores.xlsx')

if __name__ == "__main__":
    initialize_excel()
    # Example usage:
    update_scores(5, 3)  # Update scores with an example value
